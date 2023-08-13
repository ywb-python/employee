#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/6/9 12:27
# @Author  : ywb
# @Site    : 
# @File    : depart.py
# @Software: PyCharm
from django.shortcuts import render, redirect,HttpResponse

from app01 import models
from app01.utils.pagination import Pagination


def depart_list(request):
    """部门列表"""
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset, page_size=2)
    context = {"queryset": page_object.page_queryset,
               "page_string": page_object.html()}
    return render(request, 'depart_list.html', context)


def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'depart_add.html')
    title = request.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list")


def depart_delete(request):
    """删除部门"""
    nid = request.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list")


def depart_edit(request, nid):
    """修改部门"""
    raw_object = models.Department.objects.filter(id=nid).first()
    if not raw_object:
        msg = "部门信息不存在"
        return render(request, "error.html", {"msg": msg})
    if request.method == "GET":
        return render(request, 'depart_edit.html', {"raw_object": raw_object})
    title = request.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list")


def depart_multi(request):
    """批量上传excel文件"""
    file_object = request.FILES.get('exc')
    from openpyxl import load_workbook
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect("/depart/list")

